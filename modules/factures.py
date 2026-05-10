import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import date, datetime
from database import get_db
from utils import fmt_date, fmt_money

INVOICES_DIR = Path("/Users/gautiersalinier/Documents/Angels Share Marketing Limited/Angels' Share Management/Factures Commissions")
INVOICES_DIR.mkdir(parents=True, exist_ok=True)

AS_NOM     = "Angels' Share Marketing Limited"
AS_ADR1    = "15/F HONG KONG AND MACAU BLDG"
AS_ADR2    = "156-157 CONNAUGHT RD CENTRAL"
AS_ADR3    = "HONG KONG - CHINA"
AS_TEL     = "(852) 2853-7600"
AS_FAX     = "(852) 3113-1700"
AS_BANK_NOM   = "ANGELS' SHARE MARKETING LIMITED"
AS_BANK_NOM2  = "DBS Bank (Hong Kong) Limited"
AS_BANK_ACCT  = "7950161582"
AS_BANK_SWIFT = "DHBKHKHH"


def _ensure_table(db):
    try:
        db.execute("""CREATE TABLE IF NOT EXISTS commission_invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_number TEXT UNIQUE NOT NULL,
            producteur_id INTEGER,
            producteur_nom TEXT,
            prod_code TEXT,
            date_facture TEXT,
            montant_total REAL,
            devise TEXT DEFAULT 'EUR',
            description TEXT,
            contact_att TEXT,
            commandes_ids TEXT,
            fichier TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        )""")
        db.commit()
    except Exception:
        pass


def _next_invoice_number(db, prod_code, year, month):
    prefix = f"INVOICE {prod_code}{year}{month:02d}"
    existing = db.execute(
        "SELECT invoice_number FROM commission_invoices WHERE invoice_number LIKE ?",
        (f"{prefix}%",)
    ).fetchall()
    return f"{prefix}{len(existing)+1:02d}"


def generate_invoice_xlsx(invoice_number, prod, contact_att, date_facture, lignes, devise):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment

    template_path = Path(__file__).parent.parent / "INVOICE_LEDA_20260201.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"Modèle introuvable : {template_path}")

    wb = load_workbook(str(template_path))
    ws = wb.active

    font_normal = Font(name="Cambria", size=10)

    # Numéro et date
    num_part = invoice_number.replace("INVOICE ", "").replace(prod.get("code",""), "")
    try:
        ws["K9"] = int(num_part)
    except Exception:
        ws["K9"] = invoice_number
    ws["K9"].font = font_normal
    ws["K10"] = date_facture
    ws["K10"].number_format = '[$-409]d\\-mmm\\-yyyy;@'

    # Destinataire
    nom      = prod.get("nom") or ""
    adr1     = prod.get("adresse_ligne1") or prod.get("adresse") or ""
    adr2     = prod.get("adresse_ligne2") or ""
    cp_ville = " ".join(filter(None, [prod.get("code_postal"), prod.get("ville")]))
    pays_adr = prod.get("pays_adresse") or "France"

    ws["B16"] = nom.upper()
    ws["B17"] = adr1.upper() if adr1 else ""
    ws["B18"] = (cp_ville or adr2).upper() if (cp_ville or adr2) else ""
    ws["B19"] = pays_adr
    ws["C21"] = contact_att

    # Lignes
    desc_rows = [25, 26, 27, 28]
    for r in desc_rows:
        ws[f"B{r}"] = None
        ws[f"K{r}"] = None

    for i, ligne in enumerate(lignes[:4]):
        r = desc_rows[i]
        ws[f"B{r}"] = ligne["description"]
        ws[f"B{r}"].font = font_normal
        ws[f"K{r}"] = ligne["montant"]
        ws[f"K{r}"].font = font_normal
        ws[f"K{r}"].number_format = f'#,##0.00\\ "{devise}"'

    ws["K30"] = "=SUM(K25:K28)"
    ws["K30"].font = Font(name="Cambria", bold=True, size=12)
    ws["K30"].number_format = f'#,##0.00\\ "{devise}"'

    safe_num = invoice_number.replace(" ", "_").replace("/", "-")
    out_path = INVOICES_DIR / f"{safe_num}.xlsx"
    wb.save(str(out_path))
    return out_path


def render():
    st.markdown("## 📄 Factures de commissions")
    db = get_db()
    _ensure_table(db)

    tab1, tab2, tab3 = st.tabs(["➕ Nouvelle facture", "📋 Historique", "⚙️ Paramètres"])

    with tab1:
        producteurs = db.execute(
            "SELECT * FROM producteurs WHERE archived=0 ORDER BY nom"
        ).fetchall()

        if not producteurs:
            st.warning("Aucun producteur enregistré.")
            db.close()
            return

        prod_names = [p["nom"] for p in producteurs]
        sel_prod_nom = st.selectbox(
            "Producteur / débiteur *",
            ["— Sélectionner un producteur —"] + prod_names
        )
        prod = next((p for p in producteurs if p["nom"] == sel_prod_nom), None)
        if not prod:
            st.info("Sélectionnez un producteur pour continuer.")
            db.close(); return

        if prod:
            prod_dict = dict(prod)  # Convertir en dict pour éviter l'erreur .get()

            # Vérifier adresse
            has_address = bool(
                prod_dict.get("ville") or
                prod_dict.get("adresse_ligne1") or
                prod_dict.get("adresse")
            )
            if not has_address:
                st.warning(
                    f"⚠️ Aucune adresse pour {prod_dict['nom']}. "
                    "Complétez la fiche producteur d'abord."
                )

            # Afficher adresse connue
            adr_parts = [
                prod_dict.get("adresse_ligne1") or prod_dict.get("adresse") or "",
                prod_dict.get("adresse_ligne2") or "",
                " ".join(filter(None, [prod_dict.get("code_postal"), prod_dict.get("ville")])),
                prod_dict.get("pays_adresse") or "France",
            ]
            adr_str = "\n".join(p for p in adr_parts if p.strip())
            if adr_str:
                st.code(adr_str, language=None)

        st.markdown("---")

        col1, col2, col3 = st.columns(3)
        annee     = col1.number_input("Année", min_value=2020, max_value=2035,
                                       value=date.today().year)
        mois      = col2.number_input("Mois", min_value=1, max_value=12,
                                       value=date.today().month)
        date_fact = col3.date_input("Date de la facture", value=date.today())

        if prod:
            inv_num = _next_invoice_number(db, prod_dict["code"], int(annee), int(mois))
            st.info(f"📄 Numéro : **{inv_num}**")

        # Contact Att
        contacts_prod = []
        if prod:
            contacts_prod = db.execute(
                "SELECT nom, role FROM producteur_contacts WHERE producteur_id=? ORDER BY id",
                (prod_dict["id"],)
            ).fetchall()
        att_options = [""] + [f"{c['nom']} ({c['role']})" for c in contacts_prod]
        att_sel = st.selectbox("Att :", att_options)
        att_nom = att_sel.split(" (")[0] if att_sel else ""

        st.markdown("---")

        # Commandes éligibles
        st.markdown("#### Commandes à facturer")
        commandes_prod = []
        if prod:
            commandes_prod = db.execute("""
                SELECT * FROM commandes
                WHERE producteur_id=? AND archived=0
                AND comm_statut IN ('Dues','À venir')
                ORDER BY date_enlevement
            """, (prod_dict["id"],)).fetchall()

        lignes_auto = []
        if commandes_prod:
            with st.expander("🔗 Importer depuis les commandes"):
                for cmd in commandes_prod:
                    comm = (cmd["montant"] or 0) * (cmd["taux_commission"] or 0) / 100
                    checked = st.checkbox(
                        f"{cmd['proforma']} — {cmd['client_nom']} ({cmd['pays']}) "
                        f"· {cmd['comm_statut']} · Commission : {fmt_money(comm)}",
                        key=f"cmd_check_{cmd['id']}"
                    )
                    if checked:
                        # Libellé : "Commission sur" + pays + client + n° facture ou proforma
                        ref = cmd["facture_finale"] or cmd["proforma"]
                        lignes_auto.append({
                            "description": (
                                f"Commission sur {cmd['pays']} / "
                                f"{cmd['client_nom']} / {ref}"
                            ),
                            "montant": comm,
                            "cmd_id": cmd["id"]
                        })

        st.markdown("**Lignes manuelles :**")
        devise = st.selectbox("Devise", ["EUR", "USD", "HKD"], key="inv_devise")
        nb_lignes = st.number_input("Nombre de lignes", min_value=1, max_value=4, value=1)
        lignes_manuelles = []
        for i in range(int(nb_lignes)):
            c1, c2 = st.columns([3, 1])
            desc = c1.text_input(
                f"Description {i+1}",
                value="Commission sur",
                key=f"inv_desc_{i}"
            )
            mt_str = c2.text_input("Montant", placeholder="Ex: 2536.20", key=f"inv_mt_{i}")
            if desc and mt_str:
                try:
                    mt = float(mt_str.replace(",", ".").replace(" ", ""))
                    lignes_manuelles.append({"description": desc, "montant": mt})
                except ValueError:
                    st.error(f"Montant invalide ligne {i+1}")

        lignes_finales = lignes_auto if lignes_auto else lignes_manuelles
        total_preview  = sum(l["montant"] for l in lignes_finales)
        if total_preview > 0:
            st.success(f"💰 Total : **{fmt_money(total_preview, devise)}**")

        st.markdown("---")

        if st.button("📄 Générer la facture Excel", type="primary",
                     disabled=not (prod and lignes_finales)):
            if not has_address:
                st.error("Renseignez d'abord l'adresse du producteur.")
            elif not lignes_finales:
                st.error("Ajoutez au moins une ligne.")
            else:
                try:
                    out_path = generate_invoice_xlsx(
                        invoice_number=inv_num,
                        prod=prod_dict,
                        contact_att=att_nom,
                        date_facture=date_fact,
                        lignes=lignes_finales,
                        devise=devise,
                    )
                    cmd_ids = ",".join(str(l["cmd_id"]) for l in lignes_finales
                                      if "cmd_id" in l)
                    db.execute("""INSERT OR IGNORE INTO commission_invoices
                        (invoice_number, producteur_id, producteur_nom, prod_code,
                         date_facture, montant_total, devise, contact_att,
                         commandes_ids, fichier)
                        VALUES (?,?,?,?,?,?,?,?,?,?)""",
                        (inv_num, prod_dict["id"], prod_dict["nom"], prod_dict["code"],
                         date_fact.strftime("%Y-%m-%d"), total_preview, devise,
                         att_nom, cmd_ids, out_path.name))
                    db.commit()
                    st.success(f"✅ Facture **{inv_num}** générée !")
                    st.info(f"📁 Sauvegardée : `{out_path}`")
                    with open(out_path, "rb") as f:
                        st.download_button(
                            "⬇️ Télécharger la facture",
                            data=f.read(),
                            file_name=out_path.name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception as e:
                    st.error(f"Erreur : {e}")
                    import traceback
                    st.code(traceback.format_exc())

    with tab2:
        invoices = db.execute(
            "SELECT * FROM commission_invoices ORDER BY created_at DESC"
        ).fetchall()
        if not invoices:
            st.info("Aucune facture générée.")
        else:
            for inv in invoices:
                ci, cd, cdel = st.columns([7, 2, 1])
                with ci:
                    st.markdown(
                        f"**{inv['invoice_number']}** — {inv['producteur_nom']}  \n"
                        f"📅 {fmt_date(inv['date_facture'])} · "
                        f"💶 {fmt_money(inv['montant_total'], inv['devise'])} · "
                        f"Att: {inv['contact_att'] or '—'}"
                    )
                with cd:
                    fp = INVOICES_DIR / (inv["fichier"] or "")
                    if fp.exists():
                        with open(fp, "rb") as f:
                            st.download_button("⬇️ Excel", data=f.read(),
                                file_name=inv["fichier"],
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"dl_{inv['id']}")
                with cdel:
                    if st.button("🗑️", key=f"del_inv_{inv['id']}"):
                        if fp.exists(): fp.unlink()
                        db.execute("DELETE FROM commission_invoices WHERE id=?", (inv["id"],))
                        db.commit()
                        st.rerun()
                st.markdown("<hr style='margin:4px 0;border:none;border-top:1px solid #eee'>",
                            unsafe_allow_html=True)

    with tab3:
        st.markdown(f"""
| Champ | Valeur |
|---|---|
| Nom | {AS_NOM} |
| Adresse | {AS_ADR1}, {AS_ADR2}, {AS_ADR3} |
| Tél / Fax | {AS_TEL} / {AS_FAX} |
| Banque | {AS_BANK_NOM2} |
| Compte | {AS_BANK_ACCT} · SWIFT : {AS_BANK_SWIFT} |
        """)
        st.info("Pour modifier ces informations, éditez le haut de `modules/factures.py`.")

    db.close()
