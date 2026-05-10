#!/usr/bin/env python3
"""
import_prospects_sqlite.py
─────────────────────────
Importe le fichier prospects_ts_clean.xlsx dans la base SQLite angels_share.db.
À exécuter UNE FOIS depuis le terminal :

    cd "/Users/gautiersalinier/Documents/Angels Share Marketing Limited/Administratif/App Angels Share/angels_share"
    python3 import_prospects_sqlite.py

Le script est IDEMPOTENT : il ne duplique pas les prospects déjà présents
(vérification par nom + pays).
"""

import sqlite3, sys, os
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

# ── Chemins (absolus) ────────────────────────────────────────────────────────
DB_PATH    = Path("/Users/gautiersalinier/Documents/Angels Share Marketing Limited/Angels' Share Management/angels_share.db")
EXCEL_PATH = Path("/Users/gautiersalinier/Documents/Angels Share Marketing Limited/Administratif/App Angels Share/angels_share/prospects_ts_clean.xlsx")

print(f"📂 Base de données : {DB_PATH}")
print(f"📊 Fichier Excel   : {EXCEL_PATH}")

if not DB_PATH.exists():
    print(f"❌ Base de données introuvable : {DB_PATH}")
    sys.exit(1)
if not EXCEL_PATH.exists():
    print(f"❌ Fichier Excel introuvable : {EXCEL_PATH}")
    print("   Placez prospects_ts_clean.xlsx dans le dossier de l'app.")
    sys.exit(1)

# ── Connexion DB ──────────────────────────────────────────────────────────────
db = sqlite3.connect(str(DB_PATH))
db.row_factory = sqlite3.Row

# ── S'assurer que la table prospection existe et a toutes les colonnes ────────
db.execute("""CREATE TABLE IF NOT EXISTS prospection (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    civilite TEXT DEFAULT '—',
    prenom TEXT,
    nom TEXT NOT NULL,
    type TEXT DEFAULT 'Importateur / client',
    pays TEXT,
    activite TEXT,
    source TEXT,
    etape TEXT DEFAULT 'Nouveau prospect',
    contact_nom TEXT,
    contact_email TEXT,
    contact_mobile TEXT,
    contact_tel_fixe TEXT,
    contact_poste TEXT,
    contact_whatsapp TEXT,
    contact_langue TEXT,
    contact_role_email TEXT,
    producteur_interet TEXT,
    date_prochain_contact TEXT,
    raison_refus TEXT,
    notes TEXT,
    adresse TEXT,
    tel_fixe_societe TEXT,
    concurrents TEXT,
    date_fiche_source TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now')),
    archived INTEGER DEFAULT 0
)""")

db.execute("""CREATE TABLE IF NOT EXISTS prospection_interactions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    prospect_id INTEGER REFERENCES prospection(id),
    date_interaction TEXT DEFAULT (datetime('now')),
    type TEXT,
    contenu TEXT,
    notes TEXT
)""")

# Ajouter colonnes manquantes si la table existait déjà
for col, defn in [
    ("source",              "TEXT"),
    ("contact_tel_fixe",    "TEXT"),
    ("contact_poste",       "TEXT"),
    ("contact_whatsapp",    "TEXT"),
    ("contact_langue",      "TEXT"),
    ("contact_role_email",  "TEXT"),
    ("adresse",             "TEXT"),
    ("tel_fixe_societe",    "TEXT"),
    ("concurrents",         "TEXT"),
    ("date_fiche_source",   "TEXT"),
    ("civilite",            "TEXT DEFAULT '—'"),
    ("prenom",              "TEXT"),
    ("activite",            "TEXT"),
]:
    try:
        db.execute(f"ALTER TABLE prospection ADD COLUMN {col} {defn}")
    except Exception:
        pass

db.commit()

# ── Charger les prospects existants (pour éviter les doublons) ────────────────
existants = set()
for row in db.execute("SELECT nom, pays FROM prospection"):
    existants.add((str(row["nom"]).strip().lower(), str(row["pays"] or "").strip().lower()))

print(f"\n✅ {len(existants)} prospects déjà en base")

# ── Parser le fichier Excel ───────────────────────────────────────────────────
wb = load_workbook(str(EXCEL_PATH), data_only=True)
ws = wb["Clients_et_Contacts"]

inseres = 0
sautes  = 0
erreurs = 0

current_ent = None

for r in range(5, ws.max_row + 1):
    typ = str(ws.cell(r, 1).value or "").strip()

    if typ == "ENTREPRISE":
        nom     = str(ws.cell(r, 2).value or "").strip()
        if not nom:
            current_ent = None
            continue

        pays    = str(ws.cell(r, 5).value or "").strip()
        activite= str(ws.cell(r, 4).value or "").strip()
        source  = str(ws.cell(r, 28).value or "").strip()
        adresse = str(ws.cell(r, 9).value or "").strip()
        tel_soc = str(ws.cell(r, 11).value or "").strip()
        notes   = str(ws.cell(r, 19).value or "").strip()
        concu   = str(ws.cell(r, 25).value or "").strip()
        date_f  = str(ws.cell(r, 27).value or "").strip()

        current_ent = {
            "nom":      nom,
            "pays":     pays,
            "activite": activite,
            "source":   source,
            "adresse":  adresse,
            "tel_soc":  tel_soc,
            "notes":    notes,
            "concu":    concu,
            "date_f":   date_f,
            "ctcs":     [],
        }

    elif typ == "CONTACT" and current_ent:
        civ     = str(ws.cell(r, 32).value or "—").strip()
        prenom  = str(ws.cell(r, 33).value or "").strip()
        nom_c   = str(ws.cell(r, 34).value or "").strip()
        poste   = str(ws.cell(r, 35).value or "").strip()
        email   = str(ws.cell(r, 36).value or "").strip().lower()
        tel_fix = str(ws.cell(r, 37).value or "").strip()
        mobile  = str(ws.cell(r, 38).value or "").strip()
        whatsapp= str(ws.cell(r, 39).value or "").strip()
        langue  = str(ws.cell(r, 41).value or "Anglais").strip()
        role    = str(ws.cell(r, 42).value or "To").strip()

        current_ent["ctcs"].append({
            "civ": civ, "prenom": prenom, "nom": nom_c,
            "poste": poste, "email": email,
            "tel_fixe": tel_fix, "mobile": mobile,
            "whatsapp": whatsapp, "langue": langue, "role": role,
        })

    elif not typ and current_ent:
        # Fin de bloc → insérer dans la DB
        ent = current_ent
        current_ent = None

        if not ent["nom"] or not ent["ctcs"]:
            continue

        cle = (ent["nom"].lower(), ent["pays"].lower())
        if cle in existants:
            sautes += 1
            continue

        # Utiliser le premier contact comme contact principal
        # et insérer les autres dans les notes
        ctc0 = ent["ctcs"][0]
        autres_ctcs = ent["ctcs"][1:]

        # Notes complémentaires avec les autres contacts
        notes_extra = ent["notes"]
        if autres_ctcs:
            notes_extra += "\n\nAutres contacts :"
            for c in autres_ctcs:
                nom_full = f"{c['prenom']} {c['nom']}".strip()
                notes_extra += f"\n• {c['civ']} {nom_full} ({c['poste']}) — {c['email']} — {c['mobile']}"

        try:
            cur = db.execute("""INSERT INTO prospection (
                nom, type, pays, activite, source, etape,
                civilite, prenom,
                contact_nom, contact_email, contact_mobile,
                contact_tel_fixe, contact_poste,
                contact_whatsapp, contact_langue, contact_role_email,
                adresse, tel_fixe_societe, concurrents,
                date_fiche_source, notes
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                ent["nom"],
                "Importateur / client",
                ent["pays"],
                ent["activite"],
                ent["source"],
                "Nouveau prospect",
                ctc0["civ"],
                ctc0["prenom"],
                f"{ctc0['prenom']} {ctc0['nom']}".strip(),
                ctc0["email"],
                ctc0["mobile"],
                ctc0["tel_fixe"],
                ctc0["poste"],
                ctc0["whatsapp"],
                ctc0["langue"],
                ctc0["role"],
                ent["adresse"],
                ent["tel_soc"],
                ent["concu"],
                ent["date_f"],
                notes_extra.strip(),
            ))
            prospect_id = cur.lastrowid

            # Insérer les contacts supplémentaires comme interactions de type "Contact"
            for c in autres_ctcs:
                nom_full = f"{c['prenom']} {c['nom']}".strip()
                db.execute("""INSERT INTO prospection_interactions
                    (prospect_id, type, contenu, notes)
                    VALUES (?,?,?,?)""", (
                    prospect_id,
                    "Contact supplémentaire",
                    f"{c['civ']} {nom_full} — {c['poste']}",
                    f"Email: {c['email']} | Mobile: {c['mobile']} | Rôle: {c['role']}",
                ))

            existants.add(cle)
            inseres += 1

            if inseres % 100 == 0:
                db.commit()
                print(f"   … {inseres} insérés", end="\r")

        except Exception as e:
            erreurs += 1
            print(f"   ⚠️  Erreur ligne {r}: {e}")

# Dernier bloc
if current_ent and current_ent["ctcs"]:
    ent = current_ent
    cle = (ent["nom"].lower(), ent["pays"].lower())
    if cle not in existants:
        ctc0 = ent["ctcs"][0]
        try:
            db.execute("""INSERT INTO prospection (
                nom, type, pays, activite, source, etape,
                civilite, prenom, contact_nom, contact_email, contact_mobile,
                contact_tel_fixe, contact_poste, contact_whatsapp,
                contact_langue, contact_role_email,
                adresse, tel_fixe_societe, concurrents, date_fiche_source, notes
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                ent["nom"], "Importateur / client", ent["pays"],
                ent["activite"], ent["source"], "Nouveau prospect",
                ctc0["civ"], ctc0["prenom"],
                f"{ctc0['prenom']} {ctc0['nom']}".strip(),
                ctc0["email"], ctc0["mobile"], ctc0["tel_fixe"],
                ctc0["poste"], ctc0["whatsapp"], ctc0["langue"], ctc0["role"],
                ent["adresse"], ent["tel_soc"], ent["concu"], ent["date_f"],
                ent["notes"].strip(),
            ))
            inseres += 1
        except Exception as e:
            erreurs += 1

db.commit()
db.close()

print(f"""
═══════════════════════════════════════
✅ Import terminé !
   Insérés  : {inseres}
   Ignorés  : {sautes} (déjà en base)
   Erreurs  : {erreurs}
═══════════════════════════════════════
Relancez l'app Streamlit pour voir vos prospects dans le module Prospection.
""")
