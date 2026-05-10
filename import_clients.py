#!/usr/bin/env python3
"""
Script d'import / mise à jour Clients & Contacts — Angels' Share CRM
Usage: python3 import_clients.py chemin/vers/fichier.xlsx

Règles :
- Si l'entreprise existe déjà (même nom) → mise à jour
- Si le contact existe déjà (même email OU même prénom+nom dans la même entreprise) → mise à jour
- Facturation vide → copie automatique depuis livraison
- WhatsApp vide + mobile présent → OUI automatiquement
- Producteurs travaillés : lien automatique avec la table distribution
"""
import sys
import sqlite3
import re
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook

XL_PATH = sys.argv[1] if len(sys.argv) > 1 else "angels_share_import_v5_propre.xlsx"
DB_PATH = Path("/Users/gautiersalinier/Documents/Angels Share Marketing Limited/Angels' Share Management/angels_share.db")

if not Path(XL_PATH).exists():
    print(f"❌ Fichier Excel introuvable : {XL_PATH}")
    sys.exit(1)
if not DB_PATH.exists():
    print(f"❌ Base de données introuvable : {DB_PATH}")
    sys.exit(1)

conn = sqlite3.connect(str(DB_PATH))
conn.row_factory = sqlite3.Row
conn.execute("PRAGMA foreign_keys=OFF")

# ── Migrations ────────────────────────────────────────────────────────────────
for sql in [
    "ALTER TABLE entreprises ADD COLUMN producteurs_lies TEXT",
    "ALTER TABLE entreprises ADD COLUMN livraison_nom TEXT",
    "ALTER TABLE entreprises ADD COLUMN livraison_adresse TEXT",
    "ALTER TABLE entreprises ADD COLUMN livraison_pays TEXT",
    "ALTER TABLE entreprises ADD COLUMN livraison_contact TEXT",
    "ALTER TABLE entreprises ADD COLUMN facturation_nom TEXT",
    "ALTER TABLE entreprises ADD COLUMN facturation_adresse TEXT",
    "ALTER TABLE entreprises ADD COLUMN facturation_pays TEXT",
    "ALTER TABLE entreprises ADD COLUMN facturation_contact TEXT",
    "ALTER TABLE entreprises ADD COLUMN numero_tva TEXT",
    "ALTER TABLE entreprises ADD COLUMN docs_requis TEXT",
    "ALTER TABLE entreprises ADD COLUMN tarifs_envoyes INTEGER DEFAULT 0",
    "ALTER TABLE entreprises ADD COLUMN notes_facturation TEXT",
    "ALTER TABLE entreprises ADD COLUMN conditions_paiement TEXT",
    "ALTER TABLE entreprises ADD COLUMN potentiel_estime REAL",
    "ALTER TABLE entreprises ADD COLUMN saisonnalite TEXT",
    "ALTER TABLE entreprises ADD COLUMN canal_vente TEXT",
    "ALTER TABLE entreprises ADD COLUMN segment_clientele TEXT",
    "ALTER TABLE entreprises ADD COLUMN concurrents TEXT",
    "ALTER TABLE entreprises ADD COLUMN exigences_etiquetage TEXT",
    "ALTER TABLE entreprises ADD COLUMN debut_relation TEXT",
    "ALTER TABLE entreprises ADD COLUMN source_contact TEXT",
    "ALTER TABLE entreprises ADD COLUMN niveau_fidelite TEXT",
    "ALTER TABLE entreprises ADD COLUMN evaluation TEXT",
    "ALTER TABLE entreprises ADD COLUMN date_derniere_commande TEXT",
    "ALTER TABLE contacts ADD COLUMN prenom TEXT",
    "ALTER TABLE contacts ADD COLUMN civilite TEXT",
    "ALTER TABLE contacts ADD COLUMN date_naissance TEXT",
    "ALTER TABLE contacts ADD COLUMN conjoint TEXT",
    "ALTER TABLE contacts ADD COLUMN enfants TEXT",
    "ALTER TABLE contacts ADD COLUMN prefs_vins TEXT",
    "ALTER TABLE contacts ADD COLUMN prefs_cuisine TEXT",
    "ALTER TABLE contacts ADD COLUMN loisirs TEXT",
    "ALTER TABLE contacts ADD COLUMN infos_perso TEXT",
    "ALTER TABLE contacts ADD COLUMN whatsapp TEXT",
]:
    try:
        conn.execute(sql)
        conn.commit()
    except Exception:
        pass

print("✅ Migrations OK")

# ── Helpers ───────────────────────────────────────────────────────────────────
def v(ws, row, col):
    val = ws.cell(row=row, column=col).value
    if val is None: return ""
    s = str(val).strip()
    return "" if s.lower() in ("none","nan","0","0.0") else s

def date_val(ws, row, col):
    val = ws.cell(row=row, column=col).value
    if val is None: return ""
    if isinstance(val, (int, float)):
        try:
            return (datetime(1899,12,30)+timedelta(days=int(val))).strftime("%d/%m/%Y")
        except Exception: pass
    return str(val).strip()

# ── Lire Excel ────────────────────────────────────────────────────────────────
wb = load_workbook(XL_PATH, data_only=True)
ws = wb["Clients_et_Contacts"]

ins_ent = upd_ent = ins_ctc = upd_ctc = skip = 0
current_ent_id  = None
current_ent_nom = ""

for r in range(5, 700):
    typ = v(ws, r, 1)
    if not typ:
        continue

    # ══ ENTREPRISE ════════════════════════════════════════════════════════════
    if typ == "ENTREPRISE":
        nom = v(ws, r, 2)
        if not nom:
            continue

        type_ent   = v(ws, r, 3)  or "Client actif"
        activite   = v(ws, r, 4)  or "Importateur"
        pays_dest  = v(ws, r, 5)  or ""
        statut     = v(ws, r, 6)  or "Actif"
        prods_lies = v(ws, r, 7)

        # Livraison
        liv_nom  = v(ws, r, 8)
        liv_adr  = v(ws, r, 9)
        liv_pays = v(ws, r, 10)
        liv_tel  = v(ws, r, 11)

        # Facturation — si vide, copie livraison
        fact_nom  = v(ws, r, 12) or liv_nom
        fact_adr  = v(ws, r, 13) or liv_adr
        fact_pays = v(ws, r, 14) or liv_pays
        tva       = v(ws, r, 15)
        notes_f   = v(ws, r, 16)
        docs      = v(ws, r, 17)
        tarifs    = 1 if v(ws, r, 18).upper() == "OUI" else 0
        notes     = v(ws, r, 19)

        # Nouvelles colonnes
        conditions = v(ws, r, 20)
        potentiel_s= v(ws, r, 21)
        saisonnali = v(ws, r, 22)
        canal      = v(ws, r, 23)
        segment    = v(ws, r, 24)
        concurrents= v(ws, r, 25)
        exigences  = v(ws, r, 26)
        debut_rel  = date_val(ws, r, 27)
        source     = v(ws, r, 28)
        fidelite   = v(ws, r, 29)
        evaluation = v(ws, r, 30)
        ddc        = date_val(ws, r, 31)

        try:
            potentiel = float(potentiel_s.replace(" ","").replace(",",".")) if potentiel_s else None
        except Exception:
            potentiel = None

        ex = conn.execute(
            "SELECT id FROM entreprises WHERE nom=? AND archived=0", (nom,)
        ).fetchone()

        if ex:
            conn.execute("""UPDATE entreprises SET
                type=?,activite=?,pays_destination=?,statut=?,producteurs_lies=?,
                livraison_nom=?,livraison_adresse=?,livraison_pays=?,
                facturation_nom=?,facturation_adresse=?,facturation_pays=?,
                numero_tva=?,notes_facturation=?,docs_requis=?,tarifs_envoyes=?,notes=?,
                conditions_paiement=?,potentiel_estime=?,saisonnalite=?,
                canal_vente=?,segment_clientele=?,concurrents=?,
                exigences_etiquetage=?,debut_relation=?,source_contact=?,
                niveau_fidelite=?,evaluation=?,date_derniere_commande=?,archived=0
                WHERE id=?""",
                (type_ent,activite,pays_dest,statut,prods_lies,
                 liv_nom,liv_adr,liv_pays,
                 fact_nom,fact_adr,fact_pays,
                 tva,notes_f,docs,tarifs,notes,
                 conditions,potentiel,saisonnali,
                 canal,segment,concurrents,
                 exigences,debut_rel,source,
                 fidelite,evaluation,ddc,
                 ex["id"]))
            current_ent_id = ex["id"]
            upd_ent += 1
        else:
            conn.execute("""INSERT INTO entreprises
                (nom,type,activite,pays_destination,statut,producteurs_lies,
                 livraison_nom,livraison_adresse,livraison_pays,
                 facturation_nom,facturation_adresse,facturation_pays,
                 numero_tva,notes_facturation,docs_requis,tarifs_envoyes,notes,
                 conditions_paiement,potentiel_estime,saisonnalite,
                 canal_vente,segment_clientele,concurrents,
                 exigences_etiquetage,debut_relation,source_contact,
                 niveau_fidelite,evaluation,date_derniere_commande)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (nom,type_ent,activite,pays_dest,statut,prods_lies,
                 liv_nom,liv_adr,liv_pays,
                 fact_nom,fact_adr,fact_pays,
                 tva,notes_f,docs,tarifs,notes,
                 conditions,potentiel,saisonnali,
                 canal,segment,concurrents,
                 exigences,debut_rel,source,
                 fidelite,evaluation,ddc))
            current_ent_id = conn.execute(
                "SELECT id FROM entreprises WHERE nom=?", (nom,)
            ).fetchone()["id"]
            ins_ent += 1

        current_ent_nom = nom
        conn.commit()
        print(f"  {'UPD' if ex else 'NEW'} 🏢 {nom} ({pays_dest})")

    # ══ CONTACT ═══════════════════════════════════════════════════════════════
    elif typ == "CONTACT":
        if not current_ent_id:
            skip += 1
            continue

        civ    = v(ws, r, 32) or "—"
        prenom = v(ws, r, 33)
        nom_c  = v(ws, r, 34)
        if not nom_c and not prenom:
            skip += 1
            continue

        poste   = v(ws, r, 35)
        email   = v(ws, r, 36)
        tel     = v(ws, r, 37)
        mobile  = v(ws, r, 38)
        wa      = v(ws, r, 39)
        wechat  = v(ws, r, 40)
        langue  = v(ws, r, 41) or "Anglais"
        role    = v(ws, r, 42) or "To"
        ddn     = date_val(ws, r, 43)
        conjoint= v(ws, r, 44)
        enfants = v(ws, r, 45)
        p_vins  = v(ws, r, 46)
        p_cuis  = v(ws, r, 47)
        loisirs = v(ws, r, 48)
        notes_c = v(ws, r, 49)

        # Règle WhatsApp : si vide + mobile présent → OUI
        if mobile and not wa:
            wa = "OUI"

        # Chercher doublon
        ex_c = None
        if email:
            ex_c = conn.execute(
                "SELECT id FROM contacts WHERE entreprise_id=? AND email=?",
                (current_ent_id, email)).fetchone()
        if not ex_c and nom_c:
            ex_c = conn.execute(
                "SELECT id FROM contacts WHERE entreprise_id=? AND nom=? AND (prenom=? OR prenom IS NULL)",
                (current_ent_id, nom_c, prenom)).fetchone()

        if ex_c:
            conn.execute("""UPDATE contacts SET
                civilite=?,prenom=?,nom=?,position=?,email=?,
                tel_fixe=?,mobile=?,whatsapp=?,wechat=?,
                langue=?,email_role=?,date_naissance=?,
                conjoint=?,enfants=?,prefs_vins=?,prefs_cuisine=?,
                loisirs=?,notes=? WHERE id=?""",
                (civ,prenom,nom_c,poste,email,
                 tel,mobile,wa,wechat,
                 langue,role,ddn,
                 conjoint,enfants,p_vins,p_cuis,
                 loisirs,notes_c,ex_c["id"]))
            upd_ctc += 1
        else:
            conn.execute("""INSERT INTO contacts
                (entreprise_id,civilite,prenom,nom,position,email,
                 tel_fixe,mobile,whatsapp,wechat,
                 langue,email_role,date_naissance,
                 conjoint,enfants,prefs_vins,prefs_cuisine,
                 loisirs,notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (current_ent_id,civ,prenom,nom_c,poste,email,
                 tel,mobile,wa,wechat,
                 langue,role,ddn,
                 conjoint,enfants,p_vins,p_cuis,
                 loisirs,notes_c))
            ins_ctc += 1

        conn.commit()

conn.execute("PRAGMA foreign_keys=ON")
conn.commit()

print(f"\n🎉 Import terminé !")
print(f"   Entreprises : {ins_ent} créées, {upd_ent} mises à jour")
print(f"   Contacts    : {ins_ctc} créés,  {upd_ctc} mis à jour")
print(f"   Ignorés     : {skip}")
print(f"\n   Totaux en base :")
print(f"   → {conn.execute('SELECT COUNT(*) FROM entreprises WHERE archived=0').fetchone()[0]} entreprises")
print(f"   → {conn.execute('SELECT COUNT(*) FROM contacts WHERE archived=0').fetchone()[0]} contacts")
conn.close()
