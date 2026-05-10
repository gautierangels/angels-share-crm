#!/usr/bin/env python3
"""
import_prospects_sqlite_v2.py
─────────────────────────────
VERSION 2 — 1 ligne par CONTACT (pas par entreprise)
Chaque contact devient une entrée distincte dans la table prospection,
avec le nom de l'entreprise comme identifiant commun.

À exécuter depuis le terminal :
    cd ".../App Angels Share/angels_share"
    python3 import_prospects_sqlite_v2.py
"""

import sqlite3, sys
from pathlib import Path
from openpyxl import load_workbook

DB_PATH    = Path("/Users/gautiersalinier/Documents/Angels Share Marketing Limited/Angels' Share Management/angels_share.db")
EXCEL_PATH = Path("/Users/gautiersalinier/Documents/Angels Share Marketing Limited/Administratif/App Angels Share/angels_share/prospects_total_asie_clean.xlsx")

print(f"📂 Base : {DB_PATH}")
print(f"📊 Excel : {EXCEL_PATH}")

if not DB_PATH.exists():
    print(f"❌ Base introuvable"); sys.exit(1)
if not EXCEL_PATH.exists():
    print(f"❌ Excel introuvable"); sys.exit(1)

db = sqlite3.connect(str(DB_PATH))
db.row_factory = sqlite3.Row

# ── Table prospection ─────────────────────────────────────────────────────────
db.execute("""CREATE TABLE IF NOT EXISTS prospection (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    civilite TEXT DEFAULT '—',
    prenom TEXT,
    nom TEXT NOT NULL,
    type TEXT DEFAULT 'Importateur / client',
    pays TEXT,
    activite TEXT,
    source TEXT,
    etape TEXT DEFAULT 'Nouveau prospect',
    contact_nom TEXT,
    contact_email TEXT,
    contact_mobile TEXT,
    contact_tel_fixe TEXT,
    contact_poste TEXT,
    contact_whatsapp TEXT,
    contact_langue TEXT,
    contact_role_email TEXT,
    producteur_interet TEXT,
    date_prochain_contact TEXT,
    raison_refus TEXT,
    notes TEXT,
    adresse TEXT,
    tel_fixe_societe TEXT,
    concurrents TEXT,
    date_fiche_source TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now')),
    archived INTEGER DEFAULT 0
)""")

db.execute("""CREATE TABLE IF NOT EXISTS prospection_interactions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    prospect_id INTEGER REFERENCES prospection(id),
    date_interaction TEXT DEFAULT (datetime('now')),
    type TEXT, contenu TEXT, notes TEXT
)""")

# Colonnes manquantes
for col, defn in [
    ("source","TEXT"), ("contact_tel_fixe","TEXT"), ("contact_poste","TEXT"),
    ("contact_whatsapp","TEXT"), ("contact_langue","TEXT"), ("contact_role_email","TEXT"),
    ("adresse","TEXT"), ("tel_fixe_societe","TEXT"), ("concurrents","TEXT"),
    ("date_fiche_source","TEXT"), ("civilite","TEXT DEFAULT '—'"), ("prenom","TEXT"),
    ("activite","TEXT"),
]:
    try:
        db.execute(f"ALTER TABLE prospection ADD COLUMN {col} {defn}")
    except Exception: pass

db.commit()

# ── Supprimer les anciens prospects importés (reimport propre) ────────────────
print("\n🗑️  Suppression des prospects précédents (etape='Nouveau prospect')…")
db.execute("DELETE FROM prospection WHERE etape='Nouveau prospect' AND archived=0")
db.commit()
print("   Fait.")

# ── Lire le fichier Excel ─────────────────────────────────────────────────────
wb = load_workbook(str(EXCEL_PATH), data_only=True)
ws = wb["Clients_et_Contacts"]

inseres = 0
erreurs = 0
current_ent = None

for r in range(5, ws.max_row + 1):
    typ = str(ws.cell(r,1).value or "").strip()

    if typ == "ENTREPRISE":
        nom = str(ws.cell(r,2).value or "").strip()
        if not nom: current_ent = None; continue
        current_ent = {
            "nom":      nom,
            "pays":     str(ws.cell(r,5).value or "").strip(),
            "activite": str(ws.cell(r,4).value or "").strip(),
            "source":   str(ws.cell(r,28).value or "").strip(),
            "adresse":  str(ws.cell(r,9).value or "").strip(),
            "tel_soc":  str(ws.cell(r,11).value or "").strip(),
            "notes":    str(ws.cell(r,19).value or "").strip(),
            "concu":    str(ws.cell(r,25).value or "").strip(),
            "date_f":   str(ws.cell(r,27).value or "").strip(),
        }

    elif typ == "CONTACT" and current_ent:
        civ      = str(ws.cell(r,32).value or "—").strip()
        prenom   = str(ws.cell(r,33).value or "").strip()
        nom_c    = str(ws.cell(r,34).value or "").strip()
        poste    = str(ws.cell(r,35).value or "").strip()
        email    = str(ws.cell(r,36).value or "").strip().lower()
        tel_fix  = str(ws.cell(r,37).value or "").strip()
        mobile   = str(ws.cell(r,38).value or "").strip()
        whatsapp = str(ws.cell(r,39).value or "").strip()
        langue   = str(ws.cell(r,41).value or "Anglais").strip()
        role     = str(ws.cell(r,42).value or "To").strip()

        if not email: continue  # skip contacts sans email

        nom_complet = f"{prenom} {nom_c}".strip() or "—"

        # Nettoyer le nom de l'entreprise si c'est un domaine email générique
        nom_ent = current_ent["nom"]
        generiques = ['gmail','yahoo','hotmail','outlook','icloud','me.com',
                      'msn','live.','aol.','qq.com','163.com','proton',
                      'orange.','free.fr','wanadoo','laposte','126.com',
                      'rocketmail','ymail','googlemail']
        if any(g in nom_ent.lower() for g in generiques):
            nom_ent = ""  # sera affiché comme "— (indépendant)"

        try:
            db.execute("""INSERT INTO prospection (
                nom, type, pays, activite, source, etape,
                civilite, prenom,
                contact_nom, contact_email, contact_mobile,
                contact_tel_fixe, contact_poste,
                contact_whatsapp, contact_langue, contact_role_email,
                adresse, tel_fixe_societe, concurrents,
                date_fiche_source, notes
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                nom_ent or nom_complet,
                "Importateur / client",
                current_ent["pays"],
                current_ent["activite"],
                current_ent["source"],
                "Nouveau prospect",
                civ, prenom,
                nom_complet,
                email, mobile, tel_fix, poste,
                whatsapp, langue, role,
                current_ent["adresse"],
                current_ent["tel_soc"],
                current_ent["concu"],
                current_ent["date_f"],
                current_ent["notes"],
            ))
            inseres += 1
            if inseres % 200 == 0:
                db.commit()
                print(f"   … {inseres} insérés", end="\r")
        except Exception as e:
            erreurs += 1
            if erreurs <= 5: print(f"   ⚠️  Erreur L{r}: {e}")

db.commit()
db.close()

print(f"""
═══════════════════════════════════════
✅ Import v2 terminé !
   Insérés  : {inseres} contacts
   Erreurs  : {erreurs}

💡 Chaque contact est maintenant une entrée distincte.
   Relancez l'app Streamlit.
═══════════════════════════════════════
""")
